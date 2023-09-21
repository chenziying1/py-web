import openpyxl

workbook = openpyxl.Workbook()
worksheet = workbook.active

worksheet['A1'] = 'Title'
worksheet['B1'] = 'Address'
worksheet['C1'] = 'Website'
worksheet['D1'] = 'Email'
worksheet['E1'] = 'Phone'

data = [
    {'title': 'Shunshun International Education Education Agent', 'Address': 'Rm1106, A East, Jianwai SOHO, Chaoyang District, Beijing, Anhui, China', 'Website': 'www.shunshunliuxue.com', 'Email': 'liuyunxia@shunshunliuxue.com', 'phone': '该页面未显示电话信息'},
    {'title': 'ABC Education', 'Address': '123 Main Street, Anytown USA', 'Website': 'www.abc.edu', 'Email': 'info@abc.edu', 'phone': '123-456-7890'},
    # add more dictionaries as needed
]

row_num = 2

for record in data:
    worksheet.cell(row=row_num, column=1, value=record['title'])
    worksheet.cell(row=row_num, column=2, value=record['Address'])
    worksheet.cell(row=row_num, column=3, value=record['Website'])
    worksheet.cell(row=row_num, column=4, value=record['Email'])
    worksheet.cell(row=row_num, column=5, value=record['phone'])
    row_num += 1

workbook.save('data.xlsx')

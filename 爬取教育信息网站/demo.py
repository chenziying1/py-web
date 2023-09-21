# -*- coding: utf-8 -*-
# time:2023/7/23 11:13
# file index.py
# outhor:czy
# email:1060324818@qq.com
import re
import time

import chardet
import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook


def get_urls():

    urls = []
    url = 'http://www.educationagentsguide.com/china/index.htm'

    response = requests.get(url)
    soup = BeautifulSoup(response.content, 'html.parser')
    li_list = soup.find_all('li')

    title = []
    for li in li_list:
        a = li.find('a')
        if a:
            urls.append("https://www.educationagentsguide.com/china/"+a['href'])
            title.append(a.string)
    return urls,title

import openpyxl

workbook = openpyxl.Workbook()
worksheet = workbook.active

worksheet['A1'] = 'Title'
worksheet['B1'] = 'Address'
worksheet['C1'] = 'Website'
worksheet['D1'] = 'Email'
worksheet['E1'] = 'Phone'

row_num = 2

urls,title = get_urls()
data = []

for i,url in enumerate(urls):
    try:
        response = requests.get(url)

        encoding = chardet.detect(response.content)['encoding']
        html = response.content.decode(encoding)

        soup = BeautifulSoup(html, 'html.parser')
        td_list = soup.find_all('td')
        h1_text = title[i]

        td_text_list = []
        for td in td_list:
            for text in td.strings:
                target = text.strip()
                if target == '' or 'Qualifications' in target or target == "|" or "Qualification" in target or '\n\t' in target \
                        or target == "Website:" or target == "Message:" or target == "Telephone:":
                    continue
                else:
                    td_text_list.append(target)

        json_data = {'title': h1_text, 'Address': '该页面未显示地址信息', 'Website': '该页面未显示网址信息',
                     'Email': '该页面未显示邮件信息', 'phone': '该页面未显示电话信息'}
        pattern = re.compile(r'^(\d{3}-|\(\d{3}\) )?\d{3}-\d{4}$')
        pattern2 = re.compile(r'^\w+([-+.]\w+)*@\w+([-.]\w+)*\.\w+([-.]\w+)*$')
        pattern3 = re.compile(r'^https?://(?:\d{1,3}\.){3}\d{1,3}|\b(?:\d{1,3}\.){2}\d{1,3}\b')
        address_pattern = r"[A-Za-z\s\d,&]+,\s[A-Za-z\s\d,&]+,\s[A-Za-z\s\d,&]+,\s\d{6},\s[A-Za-z\s]+"

        for i in range(len(td_text_list)):
            if td_text_list[i] == "Address" or "Address" in td_text_list[i]:
                # Extracting address from the list
                address = re.sub('Education Agents', '', td_text_list[i + 1]).strip()
                json_data["Address"] = address
            elif td_text_list[i] == "Website" or "website" in td_text_list[i]:
                # Extracting website from the list
                website = td_text_list[i + 1].strip()
                json_data["Website"] = website
            elif td_text_list[i] == "Email":
                # Extracting email from the list
                email = td_text_list[i + 1].strip()
                json_data["Email"] = email
            elif "Telephone" in td_text_list[i] or "Tel" in td_text_list[i] or "Ph" in td_text_list[i]:
                # Extracting phone number from the list
                phone = td_text_list[i + 1].strip()
                json_data["phone"] = phone
            elif re.match(pattern, td_text_list[i]):
                phone = td_text_list[i].strip()
                json_data["phone"] = phone
            elif re.match(pattern2, td_text_list[i]):
                email = td_text_list[i].strip()
                json_data["Email"] = email
            elif re.match(address_pattern, td_text_list[i]):
                address = re.sub('Education Agents', '', td_text_list[i]).strip()
                json_data["Address"] = address
            elif re.match(pattern3, td_text_list[i]):
                website = td_text_list[i].strip()
                json_data["Website"] = website

        print(json_data)
        data.append(json_data)

        worksheet.cell(row=row_num, column=1, value=json_data['title'])
        worksheet.cell(row=row_num, column=2, value=json_data['Address'])
        worksheet.cell(row=row_num, column=3, value=json_data['Website'])
        worksheet.cell(row=row_num, column=4, value=json_data['Email'])
        worksheet.cell(row=row_num, column=5, value=json_data['phone'])
        row_num += 1

        workbook.save('data.xlsx')
    except:
        print("又是该死的网络！")
    finally:
        workbook.save('data.xlsx')




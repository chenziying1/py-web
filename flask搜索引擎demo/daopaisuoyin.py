import openpyxl
from collections import Counter


# 初始化倒排索引字典和相关性列表
inverted_index = {}
similarity_list = []

# 对倒排索引进行布尔搜索，并按相关性结果排序
search_term = ""
result_list = []

def quchu():
    with open("target.txt", "r", encoding="utf-8") as f:
        for i in f.readlines():
            similarity_list.append(eval(i))
    f.close()

def search_data():
    quchu()
    terms = list(search_term.split())
    result_set = set()

    for term in terms:
        if term in inverted_index:
            if len(result_set) == 0:
                result_set.update(inverted_index[term])
            else:
                # 根据布尔运算符进行交集或并集操作
                if term.lower() == 'and':
                    result_set.intersection_update(inverted_index[terms[terms.index(term) + 1]])
                elif term.lower() == 'or':
                    result_set.update(inverted_index[terms[terms.index(term) + 1]])
                elif term.lower() == 'not':
                    result_set.difference_update(inverted_index[terms[terms.index(term) + 1]])

    if len(result_set) > 0:
        # 输出搜索到的前20个相关内容
        #print(list(result_set)[:20])
        return list(result_set)[:20]
    else:
        print('数据库中暂无此内容')
        return []

def init():
    # 打开Excel文件
    wb = openpyxl.load_workbook('patentdata.xlsx')
    ws = wb.active

    # 获取所有行数和列数
    rows = ws.max_row
    cols = ws.max_column
    # print(cols) 一共3列
    # 遍历每一行数据来读取，并构建倒排索引和相关性列表
    for row in range(2, rows):
        ID = ws.cell(row=row, column=cols - 2).value
        Title = ws.cell(row=row, column=cols - 1).value
        Content = ws.cell(row=row, column=cols).value

        # 将每个词条添加到倒排索引中
        for word in (Title + Content).split():
            if word not in inverted_index:
                inverted_index[word] = set()
            inverted_index[word].add((ID, row))

        # 计算每个词条的相关性列表
        freq = Counter([word.lower() for word in (Title + Content).split()])
        similarity_list.append([word.lower(), list(freq.values())])
    with open("target.txt","w+",encoding="utf-8") as f:
        for i in similarity_list:
            f.write(i+"\n")
    f.close()


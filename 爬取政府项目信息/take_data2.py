# -*- coding: utf-8 -*-
# time:2023/7/25 20:59
# file take_data.py
# outhor:czy
# email:1060324818@qq.com

import re
import openpyxl

file_path = 'data2.txt'

workbook = openpyxl.Workbook()
worksheet = workbook.active

worksheet['A1'] = '项目名称'
worksheet['B1'] = '项目区位优势'
worksheet['C1'] = '项目基本信息'
worksheet['D1'] = '项目类型'
worksheet['E1'] = '主导业态'
worksheet['F1'] = '前期业主/实施主体'
worksheet['G1'] = '建设运营模式'
worksheet['H1'] = '项目占地面积（亩）'
worksheet['I1'] = '现状总建筑面积(m²)'
worksheet['J1'] = '更新后总建筑面积(m²)'
worksheet['K1'] = '拟保留和改造建筑面积(m²)'
worksheet['L1'] = '拟拆除建筑面积(m²)'
worksheet['M1'] = '改造中增加建筑面积(m²)'
worksheet['N1'] = '新建建筑面积(m²)'
worksheet['O1'] = '总投资额（万元）'
worksheet['P1'] = '建设地址'
worksheet['Q1'] = '完工时间'
worksheet['R1'] = '开工时间'
worksheet['S1'] = '是否入库'
worksheet['T1'] = '入库时间'
worksheet['U1'] = '项目阶段'
worksheet['V1'] = '联系人'
worksheet['W1'] = '联系方式'
worksheet['X1'] = '商务合作需求'
worksheet['Y1'] = '地块信息'
worksheet['Z1'] = '资源信息'
worksheet['AA1'] = '本级政府预算(万元)'
worksheet['AB1'] = '上级财政补助(万元)'
worksheet['AC1'] = '企业自筹(万元)'
worksheet['AD1'] = '银行贷款(万元)'
worksheet['AE1'] = '其他金融投资(万元)'
worksheet['AF1'] = '居民出资(万元)'
worksheet['AG1'] = '其他渠道(万元)'
worksheet['AH1'] = '待定(万元)'

row_num = 2

with open(file_path, 'r',encoding="utf-8") as f:
    content = f.read()
    pattern = r'\[.*?\]'  # 匹配括号中的内容
    match_list = re.findall(pattern, content)
    for match_str in match_list:
        match_str = match_str.replace('[', '').replace(']', '')  # 去掉括号
        match_list = match_str.split(',')  # 按照逗号分隔字符串

        data = {"项目名称":"","项目区位优势":"","项目基本信息":"","项目占地面积（亩）":"","现状总建筑面积(m²)":"",
                "更新后总建筑面积(m²)":"","拟保留和改造建筑面积(m²)":"","拟拆除建筑面积(m²)":"",
                "改造中增加建筑面积(m²)":"","新建建筑面积(m²)":"","总投资额（万元）":"","建设地址":"",
                "开工时间":"","完工时间":"","是否入库":"","入库时间":"","项目阶段":"","联系人":"","联系方式":"",
                "商务合作需求":"","地块信息":"","资源信息":"","本级政府预算(万元)":"","上级财政补助(万元)":"",
                "企业自筹(万元)":"","银行贷款(万元)":"","其他金融投资(万元)":"",
                "居民出资(万元)":"","其他渠道(万元)":"","项目类型":"","主导业态":"","前期业主/实施主体":"",
                "建设运营模式":""}

        for i,target in enumerate(match_list):

            if "返回" in target:
                data["项目名称"] = match_list[i+1]

            elif "项目类型"  in target:
                data["项目类型"] = match_list[i + 2]

            elif "主导业态"  in target:
                data["主导业态"] = match_list[i + 2]

            elif "前期业主/实施主体"  in target:
                data["前期业主/实施主体"] = match_list[i + 2]

            elif "建设运营模式"  in target:
                data["建设运营模式"] = match_list[i + 1]

            elif "项目区位优势"  in target:
                data["项目区位优势"] = match_list[i + 1]

            elif "更新内容"  in target:
                data["项目基本信息"] = match_list[i + 1]

            elif "项目占地面积（亩）"  in target :
                data["项目占地面积（亩）"] = match_list[i + 1]

            elif "现状总建筑面积(m²)"  in target:
                data["现状总建筑面积(m²)"] = match_list[i + 1]

            elif "更新后总建筑面积(m²)"  in target:
                data["更新后总建筑面积(m²)"] = match_list[i + 1]

            elif "拟保留和改造建筑面积(m²)"  in target:
                data["拟保留和改造建筑面积(m²)"] = match_list[i + 1]

            elif "拟拆除建筑面积(m²)"  in target:
                data["拟拆除建筑面积(m²)"] = match_list[i + 1]

            elif "改造中增加建筑面积(m²)"  in target:
                data["改造中增加建筑面积(m²)"] = match_list[i + 1]

            elif "新建建筑面积(m²)"  in target:
                data["新建建筑面积(m²)"] = match_list[i + 1]

            elif "总投资额"  in target:
                data["总投资额（万元）"] = match_list[i + 2]

            elif "建设地址"  in target:
                data["建设地址"] = match_list[i + 3]

            elif "开工时间"  in target:
                data["开工时间"] = match_list[i + 3]

            elif "完工时间"  in target:
                data["完工时间"] = match_list[i + 3]

            elif "是否入库"  in target:
                data["是否入库"] = match_list[i + 2]

            elif "入库时间" in target :
                data["入库时间"] = match_list[i]

            elif "项目阶段"  in target:
                data["项目阶段"] = match_list[i + 3]

            elif "联系人" in target :
                data["联系人"] = match_list[i+2]

            elif "联系方式"  in target:
                data["联系方式"] = match_list[i + 2]

            elif "商务合作需求" in target:
                try:
                    result = ''.join(match_list[match_list.index(" '商务合作需求'")+1:match_list.index(" '联系人'")])
                    data["商务合作需求"] = result
                except:
                    result = ''.join(match_list[match_list.index(" '商务合作需求'")+1:match_list.index(" '隐私政策'")])
                    data["商务合作需求"] = result

            elif "地块信息表" in target:
                try:
                    result = ''.join(match_list[match_list.index(" '备注说明'")+1:match_list.index(" '资金来源'")])
                    data["地块信息"] = result
                except:
                    result = ''.join(match_list[match_list.index(" '备注说明'")+1:match_list.index(" '隐私政策'")])
                    data["资源信息"] = result

            elif "资源信息表" in target:
                try:
                    result = ''.join(match_list[match_list.index(" '备注'")+1:match_list.index(" '地块信息表'")])
                    data["资源信息"] = result
                except:
                    result = ''.join(match_list[match_list.index(" '备注'")+1:match_list.index(" '隐私政策'")])
                    data["资源信息"] = result

            elif "下面是资金来源" in target:
                try:
                    data["本级政府预算(万元)"] = match_list[i + 1].replace('<td id="bjzfys">', "").replace("</td>", "")
                    data["上级财政补助(万元)"] = match_list[i + 2].replace('<td id="czbz">',"").replace("</td>","")
                    data["企业自筹(万元)"] = match_list[i + 3].replace('<td id="qyzc">', "").replace("</td>", "")
                    data["银行贷款(万元)"] = match_list[i + 4].replace('<td id="yhdk">', "").replace("</td>", "")
                    data["其他金融投资(万元)"] = match_list[i + 5].replace('<td id="qtjrtz">', "").replace("</td>", "")
                    data["居民出资(万元)"] = match_list[i + 6].replace('<td id="jmcz">', "").replace("</td>", "")
                    data["其他渠道(万元)"] = match_list[i + 7].replace('<td id="qtqd">', "").replace("</td>", "")
                    data["待定(万元)"] = match_list[i + 8].replace('<td id="dd">', "").replace("</td>", "")
                except Exception as e:
                    print(f"Error: {e}")

        '''
        worksheet['C1'] = '项目类型'
        worksheet['D1'] = '主导业态'
        worksheet['E1'] = '前期业主/实施主体'
        worksheet['F1'] = '建设运营模式'
        '''

        worksheet.cell(row=row_num, column=1, value=data['项目名称'])
        worksheet.cell(row=row_num, column=2, value=data['项目区位优势'])
        worksheet.cell(row=row_num, column=3, value=data['项目基本信息'])
        worksheet.cell(row=row_num, column=4, value=data['项目类型'])
        worksheet.cell(row=row_num, column=5, value=data['主导业态'])
        worksheet.cell(row=row_num, column=6, value=data['前期业主/实施主体'])
        worksheet.cell(row=row_num, column=7, value=data['建设运营模式'])
        worksheet.cell(row=row_num, column=8, value=data['项目占地面积（亩）'])
        worksheet.cell(row=row_num, column=9, value=data['现状总建筑面积(m²)'])
        worksheet.cell(row=row_num, column=10, value=data['更新后总建筑面积(m²)'])
        worksheet.cell(row=row_num, column=11, value=data['拟保留和改造建筑面积(m²)'])
        worksheet.cell(row=row_num, column=12, value=data['拟拆除建筑面积(m²)'])
        worksheet.cell(row=row_num, column=13, value=data['改造中增加建筑面积(m²)'])
        worksheet.cell(row=row_num, column=14, value=data['新建建筑面积(m²)'])
        worksheet.cell(row=row_num, column=15, value=data['总投资额（万元）'])
        worksheet.cell(row=row_num, column=16, value=data['建设地址'])
        worksheet.cell(row=row_num, column=17, value=data['开工时间'])
        worksheet.cell(row=row_num, column=18, value=data['完工时间'])
        worksheet.cell(row=row_num, column=19, value=data['是否入库'])
        worksheet.cell(row=row_num, column=20, value=data['入库时间'])
        worksheet.cell(row=row_num, column=21, value=data['项目阶段'])
        worksheet.cell(row=row_num, column=22, value=data['联系人'])
        worksheet.cell(row=row_num, column=23, value=data['联系方式'])
        worksheet.cell(row=row_num, column=24, value=data['商务合作需求'])
        worksheet.cell(row=row_num, column=25, value=data['地块信息'])
        worksheet.cell(row=row_num, column=26, value=data['资源信息'])
        worksheet.cell(row=row_num, column=27, value=data['本级政府预算(万元)'])
        worksheet.cell(row=row_num, column=28, value=data['上级财政补助(万元)'])
        worksheet.cell(row=row_num, column=29, value=data['企业自筹(万元)'])
        worksheet.cell(row=row_num, column=30, value=data['银行贷款(万元)'])
        worksheet.cell(row=row_num, column=31, value=data['其他金融投资(万元)'])
        worksheet.cell(row=row_num, column=32, value=data['居民出资(万元)'])
        worksheet.cell(row=row_num, column=33, value=data['其他渠道(万元)'])
        worksheet.cell(row=row_num, column=34, value=data['待定(万元)'])

        row_num += 1

        workbook.save('data2.xlsx')

        print(match_list)
        print('-----------------------------------------')


'''
备注到土地信息表
用id方式获取资金来源
用join的方式编辑地块信息与 result = ''.join(lst[lst.index('备注说明'):lst.index('资金来源')+1])
商业信息表如何搞呢？
'''

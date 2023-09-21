# -*- coding: utf-8 -*-
# time:2023/7/25 20:59
# file take_data.py
# outhor:czy
# email:1060324818@qq.com

import re

file_path = 'data.txt'
result_list = []

import openpyxl

workbook = openpyxl.Workbook()
worksheet = workbook.active

worksheet['A1'] = '项目名称'
worksheet['B1'] = '项目区位优势'
worksheet['C1'] = '项目基本信息'
worksheet['D1'] = '项目占地面积（亩）'
worksheet['E1'] = '现状总建筑面积(m²)'
worksheet['F1'] = '更新后总建筑面积(m²)'
worksheet['G1'] = '拟保留和改造建筑面积(m²)'
worksheet['H1'] = '拟拆除建筑面积(m²)'
worksheet['I1'] = '改造中增加建筑面积(m²)'
worksheet['J1'] = '新建建筑面积(m²)'
worksheet['K1'] = '总投资额（万元）'
worksheet['L1'] = '建设地址'
worksheet['N1'] = '完工时间'
worksheet['M1'] = '开工时间'
worksheet['O1'] = '是否入库'
worksheet['P1'] = '入库时间'
worksheet['Q1'] = '项目阶段'
worksheet['R1'] = '联系人'
worksheet['S1'] = '联系方式'

row_num = 2

with open(file_path, 'r',encoding="utf-8") as f:
    content = f.read()
    pattern = r'\[.*?\]'  # 匹配括号中的内容
    match_list = re.findall(pattern, content)
    for match_str in match_list:
        match_str = match_str.replace('[', '').replace(']', '')  # 去掉括号
        match_list = match_str.split(',')  # 按照逗号分隔字符串

        data = {"项目名称":"","项目区位优势":"","项目基本信息":"","项目占地面积（亩）":"","现状总建筑面积(m²)":"",
                "更新后总建筑面积(m²)":"","拟保留和改造建筑面积(m²)":"","拟拆除建筑面积(m²)":"",
                "改造中增加建筑面积(m²)":"","新建建筑面积(m²)":"","总投资额（万元）":"","建设地址":"",
                "开工时间":"","完工时间":"","是否入库":"","入库时间":"","项目阶段":"","联系人":"","联系方式":""}

        for i,target in enumerate(match_list):

            if "返回" in target:
                data["项目名称"] = match_list[i+1]

            elif "项目区位优势"  in target:
                data["项目区位优势"] = match_list[i + 1]

            elif "更新内容"  in target:
                data["项目基本信息"] = match_list[i + 1]

            elif "项目占地面积（亩）"  in target :
                data["项目占地面积（亩）"] = match_list[i + 1]

            elif "现状总建筑面积(m²)"  in target:
                data["现状总建筑面积(m²)"] = match_list[i + 1]

            elif "更新后总建筑面积(m²)"  in target:
                data["更新后总建筑面积(m²)"] = match_list[i + 1]

            elif "拟保留和改造建筑面积(m²)"  in target:
                data["拟保留和改造建筑面积(m²)"] = match_list[i + 1]

            elif "拟拆除建筑面积(m²)"  in target:
                data["拟拆除建筑面积(m²)"] = match_list[i + 1]

            elif "改造中增加建筑面积(m²)"  in target:
                data["改造中增加建筑面积(m²)"] = match_list[i + 1]

            elif "新建建筑面积(m²)"  in target:
                data["新建建筑面积(m²)"] = match_list[i + 1]

            elif "前期业主/实施主体"  in target:
                data["总投资额（万元）"] = match_list[i + 1]

            elif "建设地址"  in target:
                data["建设地址"] = match_list[i + 3]

            elif "开工时间"  in target:
                data["开工时间"] = match_list[i + 3]

            elif "完工时间"  in target:
                data["完工时间"] = match_list[i + 3]

            elif "是否入库"  in target:
                data["是否入库"] = match_list[i + 2]

            elif "入库时间" in target :
                data["入库时间"] = match_list[i]

            elif "项目阶段"  in target:
                data["项目阶段"] = match_list[i + 3]

            elif "联系人" in target :
                data["联系人"] = match_list[i+2]

            elif "联系方式"  in target:
                data["联系方式"] = match_list[i + 2]


        worksheet.cell(row=row_num, column=1, value=data['项目名称'])
        worksheet.cell(row=row_num, column=2, value=data['项目区位优势'])
        worksheet.cell(row=row_num, column=3, value=data['项目基本信息'])
        worksheet.cell(row=row_num, column=4, value=data['项目占地面积（亩）'])
        worksheet.cell(row=row_num, column=5, value=data['现状总建筑面积(m²)'])
        worksheet.cell(row=row_num, column=6, value=data['更新后总建筑面积(m²)'])
        worksheet.cell(row=row_num, column=7, value=data['拟保留和改造建筑面积(m²)'])
        worksheet.cell(row=row_num, column=8, value=data['拟拆除建筑面积(m²)'])
        worksheet.cell(row=row_num, column=9, value=data['改造中增加建筑面积(m²)'])
        worksheet.cell(row=row_num, column=10, value=data['新建建筑面积(m²)'])
        worksheet.cell(row=row_num, column=11, value=data['总投资额（万元）'])
        worksheet.cell(row=row_num, column=12, value=data['建设地址'])
        worksheet.cell(row=row_num, column=13, value=data['开工时间'])
        worksheet.cell(row=row_num, column=14, value=data['完工时间'])
        worksheet.cell(row=row_num, column=15, value=data['是否入库'])
        worksheet.cell(row=row_num, column=16, value=data['入库时间'])
        worksheet.cell(row=row_num, column=17, value=data['项目阶段'])
        worksheet.cell(row=row_num, column=18, value=data['联系人'])
        worksheet.cell(row=row_num, column=19, value=data['联系方式'])
        row_num += 1

        workbook.save('data.xlsx')

        print(match_list)
        print('-----------------------------------------')

print(result_list)

'''
备注到土地信息表
用id方式获取资金来源
用join的方式编辑地块信息与 result = ''.join(lst[lst.index('备注说明'):lst.index('资金来源')+1])
商业信息表如何搞呢？
'''

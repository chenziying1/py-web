# -*- coding: utf-8 -*-
# time:2022/12/1 12:28
# file 提取xlxl表格中关于py的部分.py
# outhor:czy
# email:1060324818@qq.com

import openpyxl
# openpyxl.load_workbook(需要打开的excel文件路径)
wb = openpyxl.load_workbook('商品列表.xlsx')


# 获取活动表对应的表对象(表对象就是Worksheet类的对象)
active_sheet = wb.active

# 根据表名获取工作簿中指定的表
sheet2 = wb['Sheet1']

# # 获取单元格对应的 Cell 对象
# a1 = sheet2['A1']      # A1 表示A列中的第一行，这儿的列号采用的是从A开始的
# print(a1)
# # 获取单元格中的内容
# content = a1.value
# print(content)      # 结果是: Rank

# 获取第二列的所有内容
row_num = sheet2.max_row     # 获取当前表中最大的行数
for row in range(1, row_num+1):
    cell = sheet2.cell(row, 2)
    cell2 = sheet2.cell(row, 3)
    if "Python" in cell.value:
        print(cell.value,cell2.value)






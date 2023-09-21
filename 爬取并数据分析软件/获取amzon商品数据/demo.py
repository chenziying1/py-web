# -*- coding: utf-8 -*-
# time:2023/8/1 11:13
# file index.py
# outhor:czy
# email:1060324818@qq.com

import json
import re

import openpyxl

def main(input_file, output_file):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    worksheet["A1"] = "更新时间"
    worksheet["B1"] = "直达的落地页"
    worksheet["C1"] = "产品评分"
    worksheet["D1"] = "目前售价"
    worksheet["E1"] = "评论数量"
    worksheet["F1"] = "主图"
    worksheet["G1"] = "商品名称"

    row_num = 2


    with open(input_file, "r",encoding="utf-8") as f:
        for line in f.readlines():
            try:
                line = line.replace("'","\"")

                data = json.loads(line.strip())

                worksheet.cell(row=row_num, column=1, value=data["更新时间"])
                worksheet.cell(row=row_num, column=2, value=data["直达的落地页"])
                worksheet.cell(row=row_num, column=3, value=data["产品评分"])
                worksheet.cell(row=row_num, column=4, value=data["目前售价"])
                worksheet.cell(row=row_num, column=5, value=data["评论数量"])
                worksheet.cell(row=row_num, column=6, value=data["主图"])
                worksheet.cell(row=row_num, column=7, value=data["商品名称"])

                row_num += 1

                workbook.save(output_file)
                print(data)
            except Exception as e:
                print(e)






if __name__ == "__main__":
    input_file_path = "amzon_Swiss_Madison.txt"
    output_file_path = "amzon_Swiss_Madison.xlsx"
    main(input_file_path, output_file_path)
    input_file_path = "amzon_American_Standard.txt"
    output_file_path = "amzon_American_Standard.xlsx"
    main(input_file_path, output_file_path)
    input_file_path = "amzon_DeerValley.txt"
    output_file_path = "amzon_DeerValley.xlsx"
    main(input_file_path, output_file_path)
    input_file_path = "amzon_Fine_Fixtures.txt"
    output_file_path = "amzon_Fine_Fixtures.xlsx"
    main(input_file_path, output_file_path)
    input_file_path = "amzon_HOROW.txt"
    output_file_path = "amzon_HOROW.xlsx"
    main(input_file_path, output_file_path)
    input_file_path = "amzon_Kohler.txt"
    output_file_path = "amzon_Kohler.xlsx"
    main(input_file_path, output_file_path)
    input_file_path = "amzon_MOHOME.txt"
    output_file_path = "amzon_MOHOME.xlsx"
    main(input_file_path, output_file_path)
    input_file_path = "amzon_Swiss_Madison.txt"
    output_file_path = "amzon_Swiss_Madison.xlsx"
    main(input_file_path, output_file_path)
    input_file_path = "amzon_TOTO.txt"
    output_file_path = "amzon_TOTO.xlsx"
    main(input_file_path, output_file_path)
